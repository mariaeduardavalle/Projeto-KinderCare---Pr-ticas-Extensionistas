import io
from datetime import date, datetime

from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agenda.models import Atendimento
from core.decorators import role_required
from pacientes.models import Paciente
from presencas.models import PresencaAula
from terapeutas.models import Terapeuta


# ── Helpers ───────────────────────────────────────────────────────────────────

def _pct_class(pct):
    if pct >= 75:
        return 'pct-good'
    if pct >= 50:
        return 'pct-warn'
    return 'pct-bad'


def _parse_periodo(request):
    hoje = date.today()
    inicio = request.GET.get('inicio') or hoje.replace(day=1).isoformat()
    fim    = request.GET.get('fim')    or hoje.isoformat()
    inicio_dt = datetime.strptime(inicio, '%Y-%m-%d').date()
    fim_dt    = datetime.strptime(fim,    '%Y-%m-%d').date()
    return inicio, fim, inicio_dt, fim_dt


def _calcular_dados(inicio_dt, fim_dt):
    dados_terapeutas = []
    for t in Terapeuta.objects.all():
        qs        = t.atendimentos.filter(data__range=[inicio_dt, fim_dt])
        total     = qs.count()
        realizados    = qs.filter(status_presenca=Atendimento.PRESENTE).count()
        faltas        = qs.filter(status_presenca=Atendimento.FALTA).count()
        justificadas  = qs.filter(status_presenca=Atendimento.FALTA_JUSTIFICADA).count()
        extras        = qs.filter(tipo=Atendimento.EXTRA).count()
        prod = round((realizados / total) * 100, 1) if total else 0
        dados_terapeutas.append({
            'nome': t.nome,
            'total_agendado': total,
            'realizados': realizados,
            'faltas': faltas,
            'faltas_justificadas': justificadas,
            'extras': extras,
            'produtividade': prod,
            'pct_class': _pct_class(prod),
        })

    dados_pacientes = []
    for p in Paciente.objects.all():
        qs        = p.atendimentos.filter(data__range=[inicio_dt, fim_dt])
        total     = qs.count()
        realizados    = qs.filter(status_presenca=Atendimento.PRESENTE).count()
        faltas        = qs.filter(status_presenca=Atendimento.FALTA).count()
        justificadas  = qs.filter(status_presenca=Atendimento.FALTA_JUSTIFICADA).count()
        extras        = qs.filter(tipo=Atendimento.EXTRA).count()
        presenca = round((realizados / total) * 100, 1) if total else 0
        dados_pacientes.append({
            'nome': p.nome,
            'responsavel': p.responsavel,
            'total_agendado': total,
            'realizados': realizados,
            'faltas': faltas,
            'faltas_justificadas': justificadas,
            'extras': extras,
            'presenca': presenca,
            'pct_class': _pct_class(presenca),
        })

    dados_aulas = []
    for p in Paciente.objects.all():
        qs        = PresencaAula.objects.filter(paciente=p, data__range=[inicio_dt, fim_dt])
        total     = qs.count()
        presentes = qs.filter(presente=True).count()
        ausentes  = total - presentes
        percentual = round((presentes / total) * 100, 1) if total else 0
        dados_aulas.append({
            'nome': p.nome,
            'total': total,
            'presentes': presentes,
            'ausentes': ausentes,
            'percentual': percentual,
            'pct_class': _pct_class(percentual),
        })

    return dados_terapeutas, dados_pacientes, dados_aulas


def _resumo(dados, campo):
    if not dados:
        return {'total': 0, 'media': '—', 'acima_75': 0, 'abaixo_50': 0}
    valores = [d[campo] for d in dados]
    return {
        'total':      len(dados),
        'media':      f"{round(sum(valores) / len(valores), 1)}%",
        'acima_75':   sum(1 for v in valores if v >= 75),
        'abaixo_50':  sum(1 for v in valores if v < 50),
    }


def _write_sheet(ws, headers, rows):
    header_fill = PatternFill(start_color='B5446E', end_color='B5446E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill(start_color='FDF0F5', end_color='FDF0F5', fill_type='solid')

    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


# ── Views ─────────────────────────────────────────────────────────────────────

@role_required('coordenacao')
def produtividade_view(request):
    inicio, fim, inicio_dt, fim_dt = _parse_periodo(request)
    aba = request.GET.get('aba', 'terapeutas')
    dados_t, dados_p, dados_a = _calcular_dados(inicio_dt, fim_dt)

    resumos = {
        'terapeutas': _resumo(dados_t, 'produtividade'),
        'pacientes':  _resumo(dados_p, 'presenca'),
        'aulas':      _resumo(dados_a, 'percentual'),
    }

    return render(request, 'relatorios/produtividade.html', {
        'dados_terapeutas': dados_t,
        'dados_pacientes':  dados_p,
        'dados_aulas':      dados_a,
        'resumo':           resumos[aba],
        'inicio':     inicio,
        'fim':        fim,
        'inicio_fmt': inicio_dt.strftime('%d/%m/%Y'),
        'fim_fmt':    fim_dt.strftime('%d/%m/%Y'),
        'aba':        aba,
    })


@role_required('coordenacao')
def exportar_excel(request):
    inicio, fim, inicio_dt, fim_dt = _parse_periodo(request)
    dados_t, dados_p, dados_a = _calcular_dados(inicio_dt, fim_dt)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Terapeutas'
    _write_sheet(ws1,
        ['Terapeuta', 'Agendados', 'Realizados', 'Faltas', 'Faltas Justificadas', 'Extras', 'Produtividade (%)'],
        [[d['nome'], d['total_agendado'], d['realizados'], d['faltas'],
          d['faltas_justificadas'], d['extras'], d['produtividade']] for d in dados_t],
    )

    ws2 = wb.create_sheet('Pacientes')
    _write_sheet(ws2,
        ['Paciente', 'Responsável', 'Agendados', 'Realizados', 'Faltas', 'Faltas Justificadas', 'Extras', 'Presença (%)'],
        [[d['nome'], d['responsavel'], d['total_agendado'], d['realizados'],
          d['faltas'], d['faltas_justificadas'], d['extras'], d['presenca']] for d in dados_p],
    )

    ws3 = wb.create_sheet('Aulas')
    _write_sheet(ws3,
        ['Paciente', 'Total de Dias', 'Presentes', 'Ausentes', 'Frequência (%)'],
        [[d['nome'], d['total'], d['presentes'], d['ausentes'], d['percentual']] for d in dados_a],
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_{inicio}_{fim}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
